"""
Author: 
Email: 

中国期货市场监控中心 - 投资者查询系统爬虫
功能：登录、验证码识别(ddddocr)、查询客户权益/风险度/保证金/当日盈亏等数据
支持：单日查询 & 批量日期范围查询（自动排除周末）
"""

import json
import os
import re
import time
from datetime import datetime, timedelta
from io import BytesIO

import ddddocr
from requests import session

# ============================================================
# 常量配置
# ============================================================
BASE_URL = "https://investorservice.cfmmc.com"
LOGIN_URL = f"{BASE_URL}/login.do"
SET_PARAM_URL = f"{BASE_URL}/customer/setParameter.do"
VERI_CODE_FLAG = 'src="/veriCode.do?t='
MAX_RETRY = 200  # 验证码最大重试次数
REQUEST_TIMEOUT = 5  # 请求超时秒数
FIELD_PATTERN = re.compile(
    r'([^<]+).*?<td class="table-normal-text" align="right">([^&]+)'
)

# 初始化 OCR（全局复用，避免重复加载模型）
_ocr = ddddocr.DdddOcr(show_ad=False)


# ============================================================
# 验证码识别
# ============================================================
def reg_img(image_data: bytes) -> str:
    """使用 ddddocr 识别验证码"""
    return _ocr.classification(image_data).strip()


# ============================================================
# HTML 工具函数
# ============================================================
def flag_filter(content: str, flag: str) -> str:
    """从 HTML 中根据 flag 提取值"""
    return content.split(flag)[1].split('"')[0]


def extract_field(content: str, label: str) -> str | None:
    """通用字段提取：根据中文标签从 HTML 表格中提取对应数值"""
    idx = content.find(label)
    if idx == -1:
        return None
    segment = content[idx:]
    match = FIELD_PATTERN.search(segment)
    if match:
        return ''.join(match.group(2).replace(',', '').strip().split())
    return None


def get_current_equity(content: str) -> float:
    """客户权益"""
    val = extract_field(content, '客户权益')
    return float(val) if val else 0.0


def get_earnest_money(content: str) -> str:
    """保证金占用"""
    return extract_field(content, '保证金占用') or ''


def get_risk_rate(content: str) -> str:
    """风险度"""
    return extract_field(content, '风险度') or ''


def get_today_income(content: str) -> str:
    """浮动盈亏（当日盈亏）"""
    return extract_field(content, '浮动盈亏') or ''


def get_tradedate_equity(content: str) -> str:
    """交易日期"""
    idx = content.find('var todayDate = ')
    if idx == -1:
        return ''
    segment = content[idx + len('var todayDate = '):]
    return segment[:segment.find(';')].strip().replace("'", "")


# ============================================================
# 核心登录逻辑
# ============================================================
def _login(ss: session, header: dict, url: str,
           user_id: str, passwd: str, max_retries: int = MAX_RETRY) -> bool:
    """
    公共的登录 + 验证码识别逻辑。

    Returns:
        True 表示登录成功（session 已保持登录状态）
    """
    res = ss.get(url, headers=header)
    content = res.content.decode()
    veri_code_url = f"{BASE_URL}/veriCode.do?t={flag_filter(content, VERI_CODE_FLAG)}"

    for i in range(max_retries):
        print(f'  第{i + 1}次尝试识别验证码')
        try:
            # 下载验证码图片
            veri_img_bytes = ss.get(veri_code_url).content
            veri_code = reg_img(veri_img_bytes)

            if veri_code and len(veri_code) == 6:
                veri_code = ''.join(filter(str.isalnum, veri_code))
                print(f'    验证码: {veri_code}')

                post_data = {
                    "showSaveCookies": '',
                    "userID": user_id,
                    "password": passwd,
                    "vericode": veri_code,
                }
                resp = ss.post(url, data=post_data, headers=header, timeout=REQUEST_TIMEOUT)
                resp_text = resp.content.decode()

                if "验证码错误" not in resp_text:
                    print('  ✓ 登录成功')
                    return True

            time.sleep(0.8)
            # 刷新验证码 URL（加时间戳）
            veri_code_url = f"{BASE_URL}/veriCode.do?t={int(time.time() * 1000)}"

        except Exception as e:
            print(f'    ✗ 异常: {e}')

    print(f'  ✗ {max_retries} 次尝试后仍失败')
    return False


def _make_header() -> dict:
    """构建请求头"""
    return {
        'Connection': 'keep-alive',
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/120.0.0.0 Safari/537.36'
        ),
    }


def _parse_result(resp_text: str, user_id: str) -> dict:
    """解析登录后的页面，返回结构化数据"""
    return {
        'userId': user_id,
        'currentEquity': get_current_equity(resp_text),
        'currentDate': get_tradedate_equity(resp_text)[1:11],
        'riskRate': get_risk_rate(resp_text),
        'todayIncome': get_today_income(resp_text),
        'earnestMoney': get_earnest_money(resp_text),
    }


# ============================================================
# 对外接口
# ============================================================
def do(user_id: str, passwd: str) -> dict | None:
    """
    登录并查询当日客户权益数据（完整版）。

    Returns:
        包含客户权益、风险度、保证金、当日盈亏等信息的字典；失败返回 None
    """
    header = _make_header()
    ss = session()

    if not _login(ss, header, LOGIN_URL, user_id, passwd):
        return None

    # 登录成功后重新请求一次以获取数据页
    res = ss.get(LOGIN_URL, headers=header)
    resp_text = res.content.decode()

    data = _parse_result(resp_text, user_id)
    print(f'  账户:{user_id}  客户权益:{data["currentEquity"]}  '
          f'交易日期:{data["currentDate"]}  风险度:{data["riskRate"]}  '
          f'保证金占用:{data["earnestMoney"]}  当日盈亏:{data["todayIncome"]}')
    return data


def do_date(user_id: str, passwd: str, trade_date: str) -> dict | None:
    """
    查询指定交易日期的客户权益数据。

    Args:
        user_id: 投资者账号
        passwd: 密码
        trade_date: 交易日期，格式 'YYYY-MM-DD'

    Returns:
        包含解析后的结构化数据 + 原始 HTML 的字典；失败返回 None
        结构: { 'parsed': {...}, 'raw_html': '...' }
    """
    header = _make_header()
    ss = session()

    if not _login(ss, header, SET_PARAM_URL, user_id, passwd):
        return None

    # 设置查询参数（指定交易日期）
    post_param_data = {
        "showSaveCookies": '',
        "byType": 'trade',
        "tradeDate": trade_date,
    }
    resp = ss.post(SET_PARAM_URL, data=post_param_data, headers=header, timeout=REQUEST_TIMEOUT)
    resp_text = resp.content.decode()

    if "验证码错误" in resp_text:
        print(f'  ✗ 查询日期 {trade_date} 时验证码失效')
        return None

    # 解析结构化数据
    current_equity = get_current_equity(resp_text)
    current_date = get_tradedate_equity(resp_text)[1:11]

    print(f'  账户:{user_id}  客户权益:{current_equity}  交易日期:{current_date}')

    return {
        # 解析后的结构化数据（用于分析/导出）
        'parsed': {
            'userId': user_id,
            'currentEquity': current_equity,
            'currentDate': current_date,
            'tradeDate': trade_date,
        },
        # 原始 HTML 页面（用于存档/重新分析）
        'raw_html': resp_text,
    }


def do_batch(user_id: str, passwd: str,
             start_date: str, end_date: str,
             output_file: str | None = None,
             save_raw: bool = True) -> list[dict]:
    """
    批量查询指定日期范围内每个交易日（排除周六日）的客户权益数据。

    输出两个部分：
      1. raw/ 目录 — 原始 HTML 存档（每个日期一个 .html 文件）
      2. 分析文件   — JSON / CSV / XLSX 结构化数据

    Args:
        user_id: 投资者账号
        passwd: 密码
        start_date: 开始日期（含），格式 'YYYY-MM-DD'
        end_date: 结束日期（含），格式 'YYYY-MM-DD'
        output_file: 输出文件路径。默认为 cfmmc_{user_id}_{start_date}_{end_date}.json
        save_raw: 是否保存原始 HTML 到 raw/ 目录，默认 True

    Returns:
        所有日期的解析后结构化数据列表
    """
    # 生成日期范围（排除周末）
    trade_dates = _generate_trade_dates(start_date, end_date)
    total = len(trade_dates)
    print(f'\n{"=" * 60}')
    print(f'批量查询: {user_id}')
    print(f'日期范围: {start_date} ~ {end_date}, 共 {total} 个交易日')
    print(f'{"=" * 60}\n')

    # 创建原始 HTML 存档目录
    raw_dir = f'raw_{user_id}_{start_date}_{end_date}'
    if save_raw:
        os.makedirs(raw_dir, exist_ok=True)

    results: list[dict] = []
    success_count = 0
    fail_count = 0

    for idx, trade_date in enumerate(trade_dates):
        print(f'[{idx + 1:2d}/{total}] 查询 {trade_date} ...')

        data = do_date(user_id, passwd, trade_date)

        if data:
            parsed = data['parsed']
            raw_html = data['raw_html']
            results.append(parsed)
            success_count += 1

            # 保存原始 HTML
            if save_raw:
                _save_raw_html(raw_dir, user_id, trade_date, raw_html)
        else:
            # 即使失败也记录一条，方便后续补查
            results.append({
                'userId': user_id,
                'currentEquity': None,
                'currentDate': '',
                'tradeDate': trade_date,
                'error': '查询失败',
            })
            fail_count += 1

        # 每次查询间隔一下，避免请求过快
        if idx < total - 1:
            time.sleep(1.5)

    # 输出统计
    print(f'\n{"=" * 60}')
    print(f'批量查询完成! 成功: {success_count}, 失败: {fail_count}, 共: {total}')
    if save_raw:
        print(f'原始HTML存档: {raw_dir}/ ({success_count} 个文件)')
    print(f'{"=" * 60}\n')

    # 保存分析文件
    if output_file is None:
        output_file = f'cfmmc_{user_id}_{start_date}_{end_date}.json'

    _save_results(results, output_file)
    print(f'分析结果已保存至: {output_file}')

    return results


# ============================================================
# 内部工具
# ============================================================
def _save_raw_html(raw_dir: str, user_id: str, trade_date: str, html_content: str) -> None:
    """保存原始 HTML 到 raw 目录，每个日期一个文件"""
    filename = os.path.join(raw_dir, f'{user_id}_{trade_date}.html')
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(html_content)


def _generate_trade_dates(start_date: str, end_date: str) -> list[str]:
    """
    生成日期范围内的所有工作日（排除周六和周日）。
    注意：不排除法定节假日，因为节假日需要动态获取日历数据。
    """
    start = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(end_date, '%Y-%m-%d')
    dates: list[str] = []
    current = start
    while current <= end:
        if current.weekday() < 5:  # 周一=0 ... 周五=4
            dates.append(current.strftime('%Y-%m-%d'))
        current += timedelta(days=1)
    return dates


def _save_results(results: list[dict], output_file: str) -> None:
    """
    将查询结果保存到本地文件。
    支持 .json / .csv / .xlsx 三种格式（根据扩展名自动选择）。
    """
    _, ext = os.path.splitext(output_file)
    ext_lower = ext.lower()

    if ext_lower == '.csv':
        _save_as_csv(results, output_file)
    elif ext_lower in ('.xlsx', '.xls'):
        _save_as_excel(results, output_file)
    else:
        # 默认 JSON 格式
        if not output_file.endswith('.json'):
            output_file += '.json'
        _save_as_json(results, output_file)


def _save_as_json(results: list[dict], filepath: str) -> None:
    """保存为 JSON 文件（带格式化，支持中文）"""
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)


def _save_as_csv(results: list[dict], filepath: str) -> None:
    """保存为 CSV 文件"""
    import csv

    if not results:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('')
        return

    # 取所有可能的 key 作为表头（保持顺序）
    fieldnames = ['userId', 'tradeDate', 'currentDate', 'currentEquity',
                  'riskRate', 'todayIncome', 'earnestMoney', 'error']

    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(results)


def _save_as_excel(results: list[dict], filepath: str) -> None:
    """保存为 Excel 文件 (.xlsx)，带表头样式和列宽自适应"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("  ⚠ 缺少 openpyxl 依赖，请执行: pip install openpyxl")
        print("  回退保存为 CSV 格式...")
        csv_path = filepath.rsplit('.', 1)[0] + '.csv'
        _save_as_csv(results, csv_path)
        return

    if not results:
        Workbook().save(filepath)
        return

    # 表头定义（中文友好名称映射）
    columns = [
        ('userId', '账户'),
        ('tradeDate', '查询日期'),
        ('currentDate', '交易日期'),
        ('currentEquity', '客户权益'),
        ('riskRate', '风险度'),
        ('todayIncome', '当日盈亏'),
        ('earnestMoney', '保证金占用'),
        ('error', '备注'),
    ]
    col_keys = [c[0] for c in columns]
    col_headers = [c[1] for c in columns]

    wb = Workbook()
    ws = wb.active
    ws.title = '客户权益数据'

    # --- 表头样式 ---
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 写入表头
    for col_idx, header_name in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写入数据行
    for row_idx, record in enumerate(results, 2):
        for col_idx, key in enumerate(col_keys, 1):
            value = record.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx >= 4 and isinstance(value, (int, float)):
                # 数值列右对齐
                cell.alignment = Alignment(horizontal='right')

    # 列宽自适应
    column_widths = [18, 14, 14, 16, 10, 14, 16, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    wb.save(filepath)


# ============================================================
# 入口
# ============================================================
if __name__ == '__main__':
    # ========== 单日查询示例 ==========
    # result = do(account, password)
    # print(result)

    # ========== 批量查询示例 ==========

    do_batch(
        user_id=account,
        passwd=password,
        start_date='2026-04-01',
        end_date='2026-05-10',
        # output_file='my_data.json',   # 可选，默认自动生成文件名
    )
